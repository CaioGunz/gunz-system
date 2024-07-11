import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class horasTrabalhadas:
    
    def __init__(self, mes, ano, dataTrabalho, cargaHoraria, horaEntrada, horaSaidaAlmoco, horaEntradaAlmoco, horaSaida, horasExtra, observacoes, qtdHoraAlmoco):
        self.mes = mes
        self.ano = ano
        self.dataTrabalho = dataTrabalho
        self.cargaHoraria = cargaHoraria
        self.horaEntrada = horaEntrada
        self.horaSaidaAlmoco = horaSaidaAlmoco
        self.horaEntradaAlmoco = horaEntradaAlmoco
        self.horaSaida = horaSaida
        self.observacoes = observacoes
        self.qtdHoraAlmoco = qtdHoraAlmoco
    
    def dicionarioDados(self):
        return {
            'Mes': self.mes,
            'Ano': self.ano,
            'Data Trabalho': self.dataTrabalho,
            'Carga Horaria': self.cargaHoraria,
            'Hora Entrada': self.horaEntrada,
            'Hora Saida Almoco': self.horaSaidaAlmoco,
            'Hora Entrada Almoco': self.horaEntradaAlmoco,
            'Hora Saida': self.horaSaida,
            'Observacoes': self.observacoes,
            'Qtd de Horas Almoco': self.qtdHoraAlmoco
        }
    
    @staticmethod
    def calculaHoras(self, horaEntrada, horaSaidaAlmoco, horaEntradaAlmoco, horaSaida):
        fmt = '$H:%M'
        horaEntrada = datetime.strptime(horaEntrada, fmt)
        horaSaidaAlmoco = datetime.strftime(horaSaidaAlmoco, fmt)
        horaEntradaAlmoco = datetime.strftime(horaEntradaAlmoco, fmt)
        horaSaida = datetime.strftime(horaSaida, fmt)
        
        tempTrabalhadoManha = (horaSaidaAlmoco - horaEntrada).total_seconds() / 3600.0
        tempTrabalhadoTarde = (horaSaida - horaEntradaAlmoco).total_seconds() / 3600.0
        totalTrabalhado = tempTrabalhadoManha + tempTrabalhadoTarde
        return totalTrabalhado
        
    @staticmethod
    def atualizaExcel(listaTrabalho, nomeArquivo):
        try:
            # Verifica se o arquivo existe
            try:
                workbook = load_workbook(nomeArquivo)
            except FileNotFoundError:
                workbook = Workbook()
                workbook.remove(workbook.active)
                
            sheet_name = 'Horas Trabalhadas'
            if sheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(sheet_name)
                # Escreve os cabeçalhos
                headers = ['Mes', 'Ano', 'Data Trabalho', 'Carga Horaria', 'Hora Entrada', 'Hora Saida Almoco', 'Hora Entrada Almoco', 'Hora Saida', 'Horas Extra Calculadas', 'Observacoes', 'Quantidade de Horas Almoco', 'Horas Trabalhadas']
                worksheet.append(headers)
            else:
                worksheet = workbook[sheet_name]

            # Verifica se os dados já existem
            existing_entries = set()
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                existing_entries.add(tuple(row))

            # Converte os dados das horas trabalhadas em um DataFrame
            data = []
            for trabalho in listaTrabalho:
                horasTrabalhadas = horasTrabalhadas.calculaHoras(trabalho.horaEntrada, trabalho.horaSaidaAlmoco, trabalho.horaEntradaAlmoco, trabalho.horaSaida)
                horasExtrasCalculadas = horasTrabalhadas - trabalho.cargaHoraria
                data.append({
                    **trabalho.dicionarioDados(),
                    'Horas Trabalhadas': horasTrabalhadas,
                    'Horas Extra Calculadas': horasExtrasCalculadas
                })

            df = pd.DataFrame(data)

            # Adiciona as novas linhas ao worksheet
            for row in dataframe_to_rows(df, index=False, header=False):
                if tuple(row) not in existing_entries:
                    worksheet.append(row)
            
            # Define o formato numérico para a coluna "Valor"
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=12, max_col=13):
                for cell in row:
                    cell.number_format = '#,##0.00'  # Formato de número com duas casas decimais
            
            # Salva o arquivo
            workbook.save(nomeArquivo)

        except Exception as e:
            print(f"Erro ao atualizar o arquivo Excel: {e}")