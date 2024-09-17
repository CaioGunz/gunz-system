import pandas as pd
import os
import uuid
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class InvestimentosSalario:
    
    def __init__(self, mes, ano, tipoInvestimento, valor):
        self.mes = mes
        self.ano = ano
        self.tipoInvestimento = tipoInvestimento
        self.valor = valor
    
    def dicionarioDados(self):
        return {
            'ID': str(uuid.uuid4()),
            'Mes': self.mes,
            'Ano': self.ano,
            'Tipo Investimento': self.tipoInvestimento,
            'Valor': float(self.valor)
        }
    
    @staticmethod
    def atualizaExcel(listaInvestimento, nomeArquivo):
        try:
            # Verifica se o arquivo existe
            try:
                workbook = load_workbook(nomeArquivo)
            except FileNotFoundError:
                workbook = Workbook()
                workbook.remove(workbook.active)
                
            sheet_name = 'Investimento e Salario'
            if sheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(sheet_name)
                # Escreve os cabeçalhos
                headers = ['ID', 'Mes', 'Ano', 'Tipo Investimento', 'Valor']
                worksheet.append(headers)
            else:
                worksheet = workbook[sheet_name]

            # Verifica se os dados já existem
            existing_entries = set()
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                existing_entries.add(tuple(row))

            # Converte os dados das investimentos em um DataFrame
            df = pd.DataFrame([investimento.dicionarioDados() for investimento in listaInvestimento])

            # Adiciona as novas linhas ao worksheet
            for row in dataframe_to_rows(df, index=False, header=False):
                if tuple(row) not in existing_entries:
                    worksheet.append(row)
            
            # Define o formato numérico para a coluna "Valor"
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=4):
                for cell in row:
                    cell.number_format = '#,##0.00'  # Formato de número com duas casas decimais
            
            # Salva o arquivo
            workbook.save(nomeArquivo)

        except Exception as e:
            print(f"Erro ao atualizar o arquivo Excel: {e}")
            
    @staticmethod
    def carregarDadosExcel(nomeArquivo):
        # Carregar dados existentes da aba "Investimento e Salario" do Excel
        try:
            if os.path.exists(nomeArquivo):
                df = pd.read_excel(nomeArquivo, sheet_name='Investimento e Salario')
                return df
            else:
                print(f"Arquivo {nomeArquivo} não encontrado.")
                return pd.DataFrame()
        except Exception as e:
            print(f"Erro ao carregar os dados: {e}")
            return pd.DataFrame()

    @staticmethod
    def salvarDadosEditados(df_editado, nomeArquivo):
        # Atualizar a aba do Excel com os dados editados
        try:
            workbook = load_workbook(nomeArquivo)
            worksheet = workbook['Investimento e Salario']

            # Limpar a aba atual e reinserir cabeçalhos
            worksheet.delete_rows(2, worksheet.max_row)
            headers = ['ID', 'Mes', 'Ano', 'Tipo Investimento', 'Valor']
            for row in dataframe_to_rows(df_editado, index=False, header=False):
                worksheet.append(row)

            workbook.save(nomeArquivo)
        except Exception as e:
            print(f"Erro ao salvar as alterações no Excel: {e}")

    @staticmethod
    def excluirDados(ids_exclusao, nomeArquivo):
        try:
            df = InvestimentosSalario.carregarDadosExcel(nomeArquivo)
            df = df[~df['ID'].isin(ids_exclusao)]  # Excluir os IDs selecionados
            InvestimentosSalario.salvarDadosEditados(df, nomeArquivo)
        except Exception as e:
            print(f"Erro ao excluir os dados: {e}")