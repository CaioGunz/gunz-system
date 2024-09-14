import pandas as pd
import os
import uuid
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class Faculdade:
    def __init__(self, nomeMateria, notaAtvd1, notaAtvd2, notaAtvd3, notaAtvd4, notaMapa, notaSGC, valorMensalidade, dataMensalidade, pago):
        self.nomeMateria = nomeMateria
        self.notaAtvd1 = notaAtvd1
        self.notaAtvd2 = notaAtvd2
        self.notaAtvd3 = notaAtvd3
        self.notaAtvd4 = notaAtvd4
        self.notaMapa = notaMapa
        self.notaSGC = notaSGC
        self.valorMensalidade = valorMensalidade
        self.dataMensalidade = dataMensalidade
        self.pago = pago
    
    def dicionarioDados(self):
        return {
            'ID': str(uuid.uuid4()),
            'Nome da Materia': self.nomeMateria,
            'Nota Atividade 1': float(self.notaAtvd1),
            'Nota Atividade 2': float(self.notaAtvd2),
            'Nota Atividade 3': float(self.notaAtvd3),
            'Nota Atividade 4': float(self.notaAtvd4),
            'Nota Mapa': float(self.notaMapa),
            'Nota SGC': float(self.notaSGC),
            'Valor Mensalidade': float(self.valorMensalidade),
            'Data Mensalidade': self.dataMensalidade,
            'Pago': self.pago
        }
    
    @staticmethod
    def atualizaExcel(listaFaculdades, nomeArquivo):
        try:
            # Verifica se o arquivo existe e se é válido
            if os.path.exists(nomeArquivo):
                try:
                    workbook = load_workbook(nomeArquivo)
                except Exception as e:
                    print(f"Arquivo corrompido, recriando... Erro: {e}")
                    os.remove(nomeArquivo)
                    workbook = Workbook()
                    workbook.remove(workbook.active)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
                
            sheet_name = 'Faculdade'
            if sheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(sheet_name)
                # Escreve os cabeçalhos
                headers = ['ID', 
                           'Nome Materia', 
                           'Nota Atividade 1', 
                           'Nota Atividade 2', 
                           'Nota Atividade 3', 
                           'Nota Atividade 4', 
                           'Nota MAPA', 
                           'Nota SGC', 
                           'Valor Mensalidade', 
                           'Data Mensalidade', 
                           'Pago']
                worksheet.append(headers)
            else:
                worksheet = workbook[sheet_name]

            # Verifica se os dados já existem
            existing_entries = set()
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                existing_entries.add(tuple(row))

            # Converte os dados das faculdades em um DataFrame
            df = pd.DataFrame([faculdade.dicionarioDados() for faculdade in listaFaculdades])

            # Adiciona as novas linhas ao worksheet
            for row in dataframe_to_rows(df, index=False, header=False):
                if tuple(row) not in existing_entries:
                    worksheet.append(row)
            
            # Formato para data
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=10, max_col=10):
                for cell in row:
                    cell.number_format = 'dd/mm/yyyy'
            
            # Salva o arquivo
            workbook.save(nomeArquivo)

        except Exception as e:
            print(f"Erro ao atualizar o arquivo Excel: {e}")

    @staticmethod
    def carregarDadosExcel(nomeArquivo):
        # Carregar dados existentes da aba "Faculdade" do Excel
        try:
            if os.path.exists(nomeArquivo):
                df = pd.read_excel(nomeArquivo, sheet_name='Faculdade')
                return df
            else:
                print(f"Arquivo {nomeArquivo} não encontrado.")
                return pd.DataFrame()
        except Exception as e:
            print(f"Erro ao carregar os dados: {e}")
            return pd.DataFrame()

    @staticmethod
    def salvarDadosEditados(df_editado, nomeArquivo):
        # Atualizar a aba do Excel com os dados editados
        try:
            workbook = load_workbook(nomeArquivo)
            worksheet = workbook['Faculdade']

            # Limpar a aba atual e reinserir cabeçalhos
            worksheet.delete_rows(2, worksheet.max_row)
            headers = ['ID', 'Nome Materia', 'Nota Atividade 1', 'Nota Atividade 2', 'Nota Atividade 3', 'Nota Atividade 4', 'Nota MAPA', 'Nota SGC', 'Valor Mensalidade', 'Data Mensalidade', 'Pago']
            for row in dataframe_to_rows(df_editado, index=False, header=False):
                worksheet.append(row)

            workbook.save(nomeArquivo)
        except Exception as e:
            print(f"Erro ao salvar as alterações no Excel: {e}")

    @staticmethod
    def excluirDados(ids_exclusao, nomeArquivo):
        try:
            df = Faculdade.carregarDadosExcel(nomeArquivo)
            df = df[~df['ID'].isin(ids_exclusao)]  # Excluir os IDs selecionados
            Faculdade.salvarDadosEditados(df, nomeArquivo)
        except Exception as e:
            print(f"Erro ao excluir os dados: {e}")