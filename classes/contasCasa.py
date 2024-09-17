import pandas as pd
import uuid
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class contasDeCasa:
    def __init__(self, nomeContas, valor, dataVencimento, dataPagamento, pago, observacao):
        self.nomeContas = nomeContas
        self.valor = valor
        self.dataVencimento = dataVencimento
        self.dataPagamento = dataPagamento
        self.pago = pago
        self.observacao = observacao
    
    def dicionarioDados(self):
        return {
            'ID': str(uuid.uuid4()),
            'Nome Contas': self.nomeContas,
            'Valor': self.valor,
            'Data Vencimento': self.dataVencimento.strftime('%d/%m/%Y'),
            'Data Pagamento': self.dataPagamento.strftime('%d/%m/%Y'),
            'Pago': self.pago,
            'Observacao': self.observacao
        }
        
    @staticmethod
    def atualizaExcel(listaContas, nomeArquivo):
        try:
            # Verifica se o arquivo existe e se é válido
            if os.path.exists(nomeArquivo):
                try:
                    workbook = load_workbook(nomeArquivo)
                except Exception as e:
                    print(f"Arquivo corrompido, recriando... Erro: {e}")
                    os.remove(nomeArquivo)
                    workbook = Workbook()
                    workbook.remove(workbook.active)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
                
            sheet_name = 'Contas de Casa'
            if sheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(sheet_name)
                # Escreve os cabeçalhos
                headers = ['ID', 
                           'Nome Contas', 
                           'Valor', 
                           'Data Vencimento', 
                           'Data Pagamento', 
                           'Pago', 
                           'Observacao']
                worksheet.append(headers)
            else:
                worksheet = workbook[sheet_name]

            # Verifica se os dados já existem
            existing_entries = set()
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                existing_entries.add(row[0])

            # Converte os dados das contas em um DataFrame
            df = pd.DataFrame([conta.dicionarioDados() for conta in listaContas])

            # Adiciona as novas linhas ao worksheet
            for row in dataframe_to_rows(df, index=False, header=False):
                if row[0] not in existing_entries: # Checa se o ID ja existe
                    worksheet.append(row)
            
            # Formatação das células de data
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=5):
                for cell in row:
                    cell.number_format = 'DD/MM/YYYY'
            
            # Salva o arquivo
            workbook.save(nomeArquivo)

        except Exception as e:
            print(f"Erro ao atualizar o arquivo Excel: {e}")

    @staticmethod
    def carregarDadosExcel(nomeArquivo):
        # Carregar dados existentes da aba "Contas de Casa" do Excel
        try:
            if os.path.exists(nomeArquivo):
                df = pd.read_excel(nomeArquivo, sheet_name='Contas de Casa')
                return df
            else:
                print(f"Arquivo {nomeArquivo} não encontrado.")
                return pd.DataFrame()
        except Exception as e:
            print(f"Erro ao carregar os dados: {e}")
            return pd.DataFrame()

    @staticmethod
    def salvarDadosEditados(df_editado, nomeArquivo):
        # Atualizar a aba do Excel com os dados editados
        try:
            workbook = load_workbook(nomeArquivo)
            worksheet = workbook['Contas de Casa']

            # Limpar a aba atual e reinserir cabeçalhos
            worksheet.delete_rows(2, worksheet.max_row)
            headers = ['ID', 
                       'Nome Contas', 
                       'Valor', 
                       'Data Vencimento', 
                       'Data Pagamento', 
                       'Pago', 
                       'Observacao']
            for row in dataframe_to_rows(df_editado, index=False, header=False):
                worksheet.append(row)

            workbook.save(nomeArquivo)
        except Exception as e:
            print(f"Erro ao salvar as alterações no Excel: {e}")

    @staticmethod
    def excluirDados(ids_exclusao, nomeArquivo):
        try:
            df = contasDeCasa.carregarDadosExcel(nomeArquivo)
            df = df[~df['ID'].isin(ids_exclusao)]  # Excluir os IDs selecionados
            contasDeCasa.salvarDadosEditados(df, nomeArquivo)
        except Exception as e:
            print(f"Erro ao excluir os dados: {e}")