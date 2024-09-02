import pandas as pd
import os
import uuid
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class anotacaoContas:
    def __init__(self, mes, ano, categoria, descricaoProduto, valor, dataCompra, pago, observacao):
        self.mes = mes
        self.ano = ano
        self.categoria = categoria
        self.descricaoProduto = descricaoProduto
        self.valor = valor
        self.dataCompra = dataCompra
        self.pago = pago
        self.observacao = observacao

    def dicionarioDados(self):
        return {
            "ID": str(uuid.uuid4()), # Gera um ID unico para cada registro de dados
            "Mes": self.mes,
            "Ano": self.ano,
            'Categoria Conta': self.categoria,
            'Descricao': self.descricaoProduto,
            'Valor': self.valor,
            'Data Compra': self.dataCompra.strftime('%d/%m/%Y'),
            'Pago': self.pago,
            'Observacao': self.observacao
        }
    
    @staticmethod
    def atualizaExcel(listaContas, nomeArquivo):
        try:
            # Verifica se o arquivo existe e se é válido
            if os.path.exists(nomeArquivo):
                try:
                    workbook = load_workbook(nomeArquivo)
                except Exception as e:
                    print(f"Arquivo corrompido, recriando... Erro: {e}")
                    os.remove(nomeArquivo)
                    workbook = Workbook()
                    workbook.remove(workbook.active)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
            
            sheet_name = 'Anotacao Contas'
            if sheet_name not in workbook.sheetnames:
                worksheet = workbook.create_sheet(sheet_name)
                # Escreve os cabeçalhos
                headers = ['ID', 
                           'Mes', 
                           'Ano', 
                           'Categoria', 
                           'Descricao', 
                           'Valor', 
                           'Data Compra', 
                           'Pago', 
                           'Observacao']
                worksheet.append(headers)
            else:
                worksheet = workbook[sheet_name]

            # Verifica se os dados já existem
            existing_entries = set()
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                existing_entries.add(row[0]) # Salva o ID da linha existente

            # Converte os dados das contas em um DataFrame
            df = pd.DataFrame([conta.dicionarioDados() for conta in listaContas])
            
            # Adiciona as novas linhas ao worksheet, checando se o ID já existe
            for row in dataframe_to_rows(df, index=False, header=False):
                if row[0] not in existing_entries: # Checa se o ID ja existe
                    worksheet.append(row)
                
            # Formato para data
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=7, max_col=7):
                for cell in row:
                    cell.number_format = 'dd/mm/yyyy'
            
            # Salva o arquivo
            workbook.save(nomeArquivo)

        except Exception as e:
            print(f"Erro ao atualizar o arquivo Excel: {e}")
            
    @staticmethod
    def carregarDadosExcel(nomeArquivo):
        # Carregar dados existentes da aba "Anotacao Contas" do Excel
        try:
            if os.path.exists(nomeArquivo):
                df = pd.read_excel(nomeArquivo, sheet_name='Anotacao Contas')
                return df
            else:
                print(f"Arquivo {nomeArquivo} não encontrado.")
                return pd.DataFrame()
        except Exception as e:
            print(f"Erro ao carregar os dados: {e}")
            return pd.DataFrame()

    @staticmethod
    def salvarDadosEditados(df_editado, nomeArquivo):
        # Atualizar a aba do Excel com os dados editados
        try:
            workbook = load_workbook(nomeArquivo)
            worksheet = workbook['Anotacao Contas']

            # Limpar a aba atual e reinserir cabeçalhos
            worksheet.delete_rows(2, worksheet.max_row)
            headers = ['ID', 'mes', 'ano', 'categoria', 'descricaoProduto', 'valor', 'dataCompra', 'pago', 'observacao']
            for row in dataframe_to_rows(df_editado, index=False, header=False):
                worksheet.append(row)

            workbook.save(nomeArquivo)
        except Exception as e:
            print(f"Erro ao salvar as alterações no Excel: {e}")

    @staticmethod
    def excluirDados(ids_exclusao, nomeArquivo):
        try:
            df = anotacaoContas.carregarDadosExcel(nomeArquivo)
            df = df[~df['ID'].isin(ids_exclusao)]  # Excluir os IDs selecionados
            anotacaoContas.salvarDadosEditados(df, nomeArquivo)
        except Exception as e:
            print(f"Erro ao excluir os dados: {e}")